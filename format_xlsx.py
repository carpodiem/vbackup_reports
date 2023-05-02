import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font

def format_excel_file(input_file, output_file):
    # Read the input file
    df = pd.read_excel(input_file)

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write DataFrame to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Apply formatting here
    header_fill = PatternFill(start_color="0000FFFF", end_color="0000FFFF", fill_type="solid")
    header_font = Font(bold=True, size=12)

    # Formatting cells
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in ws.iter_rows(min_row=2, max_col=6, max_row=ws.max_row):
        for cell in row:
            if cell.column == 5:  # Current Backup Size GB
                cell.number_format = "#,##0.00"

    # Save the formatted Excel file
    wb.save(output_file)